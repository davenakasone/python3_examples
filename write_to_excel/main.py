"""
you can write to excel file  it is a lot easier to let python work for you when tasks start getting big
the programs can create and edit spreadsheets (.xlsx)

nothing really happens until you call save()   fyi, always keep original in tact

Font() takes:
name =     'Calibri', 'Times New Roman',.....
size =         any integer
bold =          True to turn on
italic =        True to turn on
"""
import openpyxl
from openpyxl.styles import Font

wb_obj = openpyxl.Workbook() # a new workbook is made
print(wb_obj.sheetnames)     # see what sheets it has   just has one
sheet_obj = wb_obj.active    # take it to a worksheet obj
print(sheet_obj.title)       # check the title, should be the same
sheet_obj.title = 'spe'      # change the title   spam, bacon, eggs
print(sheet_obj.title)       # check
print(wb_obj.sheetnames)     # check
#wb_obj.save('example.xlsx')  # appears to overwrite   index a sequence if that is a problem use copy if working with orignal

print('\n')
wb_obj.create_sheet()        # a sheet is made
print(wb_obj.sheetnames)     # check
wb_obj.create_sheet(index = 0, title = 'first sheet')   # create a sheet at first index
print(wb_obj.sheetnames)     # check
wb_obj.create_sheet(title = 'last sheet')        # not specified, goes to last index
print(wb_obj.sheetnames)     # check

print('\n') # delete any sheet with del
del wb_obj['last sheet']
del wb_obj['Sheet']
del wb_obj['first sheet']
print(wb_obj.sheetnames)     # check

print('\n') # writing values to cells just like dictionary      
sheet_obj = wb_obj['spe'] # take sheet to write to
sheet_obj['A1'] = 'takbir'    # also can use coordinates if have as string
print(sheet_obj['A1'].value)  # check

print('\n') # make a custom font
italic24 = Font(size = 24, italic = True)  # make a font object
sheet_obj['A1'].font = italic24            # apply to cell
#wb_obj.save('example.xlsx')                # save change

print('\n') # configure for formulas
sheet_obj['C1'] = 1
print(sheet_obj['C1'].value)
sheet_obj['C2'] = 2
print(sheet_obj['C2'].value)

sheet_obj['C3'] = 3 # wtf?
print(sheet_obj['C4'].value)
sheet_obj['C4'] = '=SUM(C1:C3)'
print(sheet_obj['C4'].value)
sheet_obj['C4'].font = italic24
wb_obj.save('example.xlsx') 